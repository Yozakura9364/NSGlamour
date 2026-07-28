import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Dict, List, Mapping, NamedTuple, Optional, Set, Tuple, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


OUTPUT_HEADERS = (
    "第一列（原始）",
    "第二列（中文）",
    "第三列（中文）",
    "第四列（日文）",
    "第五列（日文）",
    "第六列（英文）",
    "第七列（英文）",
    "第八列（日英混合）",
)


class LocalizationDatabase(NamedTuple):
    item_names: Dict[str, Dict[str, str]]
    stain_names: Dict[str, Dict[str, str]]


class ExportResult(NamedTuple):
    row_count: int
    missing_items: Tuple[str, ...]
    missing_dyes: Tuple[str, ...]


def normalize_lookup_name(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    return re.sub(r"\s+", " ", text).strip().casefold()


def clean_names(value: object) -> Dict[str, str]:
    if not isinstance(value, Mapping):
        return {}
    return {
        str(locale): str(name).strip()
        for locale, name in value.items()
        if name is not None and str(name).strip()
    }


def add_item_records(index: Dict[str, Dict[str, str]], records: object) -> None:
    values = records.values() if isinstance(records, Mapping) else records
    if not isinstance(values, (list, tuple)) and not hasattr(values, "__iter__"):
        return

    for record in values:
        if not isinstance(record, Mapping):
            continue
        names = clean_names(record.get("names"))
        source_name = names.get("zh") or str(record.get("name") or "").strip()
        key = normalize_lookup_name(source_name)
        if key and names:
            index.setdefault(key, names)


def build_stain_index(mapping: Mapping[str, object]) -> Dict[str, Dict[str, str]]:
    stains_by_locale = mapping.get("stains_by_locale")
    if not isinstance(stains_by_locale, Mapping):
        return {}

    zh_stains = stains_by_locale.get("zh")
    if not isinstance(zh_stains, Mapping):
        return {}

    index: Dict[str, Dict[str, str]] = {}
    for stain_id, source_name in zh_stains.items():
        names = {
            str(locale): str(locale_stains.get(str(stain_id))).strip()
            for locale, locale_stains in stains_by_locale.items()
            if isinstance(locale_stains, Mapping)
            and locale_stains.get(str(stain_id)) is not None
            and str(locale_stains.get(str(stain_id))).strip()
        }
        key = normalize_lookup_name(source_name)
        if key and names:
            index.setdefault(key, names)

    index[normalize_lookup_name("无")] = {
        "zh": "无",
        "ja": "無し",
        "en": "None",
        "ko": "없음",
        "tc": "無",
        "fr": "Aucune",
        "de": "Keine",
    }
    return index


def load_localization_database(mapping_path: Union[Path, str]) -> LocalizationDatabase:
    path = Path(mapping_path)
    with path.open("r", encoding="utf-8") as handle:
        mapping = json.load(handle)
    if not isinstance(mapping, Mapping):
        raise ValueError("装备映射文件结构无效")

    item_names: Dict[str, Dict[str, str]] = {}
    add_item_records(item_names, mapping.get("items") or [])
    add_item_records(item_names, mapping.get("glasses") or {})
    add_item_records(item_names, mapping.get("ornaments") or {})
    return LocalizationDatabase(item_names=item_names, stain_names=build_stain_index(mapping))


def excel_safe_text(value: str) -> str:
    return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value


def translate_term(
    value: str,
    index: Mapping[str, Mapping[str, str]],
    target_locale: str,
    missing: Set[str],
) -> str:
    source = value.strip()
    if not source:
        return ""
    names = index.get(normalize_lookup_name(source))
    if not names:
        missing.add(source)
        return excel_safe_text(source)
    return excel_safe_text(str(names.get(target_locale) or names.get("zh") or source).strip())


def translate_line(
    line: str,
    index: Mapping[str, Mapping[str, str]],
    target_locale: str,
    missing: Set[str],
) -> str:
    source = line.strip()
    if not source:
        return ""

    separator_match = re.search(r"[|｜/／]", source)
    if not separator_match:
        return translate_term(source, index, target_locale, missing)

    separator = separator_match.group(0)
    parts = [part.strip() for part in re.split(r"[|｜/／]", source) if part.strip()]
    return f" {separator} ".join(
        translate_term(part, index, target_locale, missing) for part in parts
    )


def translate_cell(
    value: object,
    index: Mapping[str, Mapping[str, str]],
    target_locale: str,
    missing: Set[str],
) -> str:
    if value is None:
        return ""
    lines = [line for line in str(value).replace("\r\n", "\n").split("\n") if line.strip()]
    return "\n".join(translate_line(line, index, target_locale, missing) for line in lines)


def translate_item_cell(
    value: object,
    database: LocalizationDatabase,
    target_locale: str,
    missing: Set[str],
) -> str:
    return translate_cell(value, database.item_names, target_locale, missing)


def translate_dye_cell(
    value: object,
    database: LocalizationDatabase,
    target_locale: str,
    missing: Set[str],
) -> str:
    return translate_cell(value, database.stain_names, target_locale, missing)


def generate_ja_en_mixed(
    value: object,
    database: LocalizationDatabase,
    missing: Set[str],
) -> str:
    if value is None:
        return ""
    entries = [line for line in str(value).replace("\r\n", "\n").split("\n") if line.strip()]
    mixed = []
    for entry in entries:
        ja = translate_line(entry, database.item_names, "ja", missing)
        en = translate_line(entry, database.item_names, "en", missing)
        mixed.append(f"{ja}\n{en}")
    return "\n\n".join(mixed)


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def export_workbook(
    input_path: Union[Path, str],
    output_path: Union[Path, str],
    mapping_path: Union[Path, str],
) -> ExportResult:
    source = Path(input_path).resolve()
    target = Path(output_path).resolve()
    if source == target:
        raise ValueError("输入文件和输出文件不能相同")
    if not source.is_file():
        raise FileNotFoundError(f"找不到输入文件：{source}")

    database = load_localization_database(mapping_path)
    missing_items: Set[str] = set()
    missing_dyes: Set[str] = set()
    input_workbook = load_workbook(source, read_only=True, data_only=False)
    input_sheet = input_workbook.active

    rows = []
    try:
        for values in input_sheet.iter_rows(min_row=2, values_only=True):
            original = cell_text(values[0] if len(values) > 0 else None)
            items_zh = cell_text(values[1] if len(values) > 1 else None)
            dyes_zh = cell_text(values[2] if len(values) > 2 else None)
            if not original and not items_zh and not dyes_zh:
                continue
            rows.append(
                (
                    excel_safe_text(original),
                    excel_safe_text(items_zh),
                    excel_safe_text(dyes_zh),
                    translate_item_cell(items_zh, database, "ja", missing_items),
                    translate_dye_cell(dyes_zh, database, "ja", missing_dyes),
                    translate_item_cell(items_zh, database, "en", missing_items),
                    translate_dye_cell(dyes_zh, database, "en", missing_dyes),
                    generate_ja_en_mixed(items_zh, database, missing_items),
                )
            )
    finally:
        input_workbook.close()

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "导出结果"
    output_sheet.append(OUTPUT_HEADERS)
    for row in rows:
        output_sheet.append(row)

    for column in range(1, 9):
        output_sheet.column_dimensions[output_sheet.cell(1, column).column_letter].width = 40
    for row in output_sheet.iter_rows(min_col=4, max_col=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    target.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(target)
    output_workbook.close()
    return ExportResult(
        row_count=len(rows),
        missing_items=tuple(sorted(missing_items)),
        missing_dyes=tuple(sorted(missing_dyes)),
    )


def default_mapping_path() -> Path:
    return Path(__file__).resolve().parents[2] / "data" / "item_model_mapping.json"


def select_input_file() -> str:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    try:
        return filedialog.askopenfilename(
            title="选择配装 Excel",
            filetypes=(("Excel 工作簿", "*.xlsx"), ("所有文件", "*.*")),
        )
    finally:
        root.destroy()


def show_message(title: str, message: str, error: bool = False) -> None:
    try:
        import ctypes

        ctypes.windll.user32.MessageBoxW(0, message, title, 0x10 if error else 0x40)
    except Exception:
        print(f"{title}: {message}")


def build_result_message(output_path: Path, result: ExportResult) -> str:
    lines = [f"导出完成，共处理 {result.row_count} 行。", "", str(output_path)]
    if result.missing_items:
        lines.extend(("", f"未匹配装备：{len(result.missing_items)} 个", "、".join(result.missing_items[:20])))
    if result.missing_dyes:
        lines.extend(("", f"未匹配染剂：{len(result.missing_dyes)} 个", "、".join(result.missing_dyes[:20])))
    return "\n".join(lines)


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="使用 NSGlamour 数据生成中日英配装 Excel")
    parser.add_argument("input", nargs="?", help="输入 .xlsx；省略时打开文件选择器")
    parser.add_argument("--output", help="输出 .xlsx；默认写到输入文件旁的 导出结果.xlsx")
    parser.add_argument("--mapping", default=str(default_mapping_path()), help="NSGlamour 映射 JSON")
    parser.add_argument("--no-dialog", action="store_true", help="仅使用控制台输出")
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    input_value = args.input or select_input_file()
    if not input_value:
        return 0

    input_path = Path(input_value)
    output_path = Path(args.output) if args.output else input_path.with_name("导出结果.xlsx")
    try:
        result = export_workbook(input_path, output_path, args.mapping)
    except Exception as exc:
        if args.no_dialog:
            print(f"导出失败：{exc}", file=sys.stderr)
        else:
            show_message("配装批量导出 - 错误", f"导出失败：\n{exc}", error=True)
        return 1

    message = build_result_message(output_path.resolve(), result)
    if args.no_dialog:
        print(message)
    else:
        show_message("配装批量导出", message)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
