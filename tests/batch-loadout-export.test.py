import importlib.util
import json
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook


tool_path = (
    Path(__file__).resolve().parents[1]
    / "tools"
    / "batch-loadout-export"
    / "batch_loadout_export.py"
)
spec = importlib.util.spec_from_file_location("batch_loadout_export", tool_path)
tool = importlib.util.module_from_spec(spec)
spec.loader.exec_module(tool)


def write_mapping(path: Path) -> None:
    mapping = {
        "metadata": {"locales": ["zh", "en", "ja"]},
        "items": [
            {
                "key": 1,
                "names": {
                    "zh": "旦衣礼帽",
                    "ja": "ボルトライズ・ハット",
                    "en": "Boltrise Hat",
                },
            },
            {
                "key": 2,
                "names": {
                    "zh": "山间少女长裙",
                    "ja": "ディアンドルロングスカート",
                    "en": "Dirndl's Long Skirt",
                },
            },
        ],
        "glasses": {},
        "ornaments": {},
        "stains_by_locale": {
            "zh": {"0": "无染色", "1": "煤玉黑"},
            "ja": {"0": "染色無し", "1": "ジェットブラック"},
            "en": {"0": "No Color", "1": "Jet Black"},
        },
    }
    path.write_text(json.dumps(mapping, ensure_ascii=False), encoding="utf-8")


def write_input(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None, None, None])
    sheet.append(
        [
            "测试配装",
            "旦衣礼帽\n山间少女长裙\n=1+1",
            "煤玉黑 | 无染色\n无",
        ]
    )
    workbook.save(path)
    workbook.close()


with tempfile.TemporaryDirectory() as temp_dir:
    root = Path(temp_dir)
    mapping_path = root / "item_model_mapping.json"
    input_path = root / "配装.xlsx"
    output_path = root / "导出结果.xlsx"
    write_mapping(mapping_path)
    write_input(input_path)

    database = tool.load_localization_database(mapping_path)
    missing_items = set()
    missing_dyes = set()
    assert tool.translate_item_cell("旦衣礼帽\n未知装备", database, "ja", missing_items) == (
        "ボルトライズ・ハット\n未知装备"
    )
    assert missing_items == {"未知装备"}
    assert tool.translate_dye_cell("煤玉黑 | 无染色\n无", database, "en", missing_dyes) == (
        "Jet Black | No Color\nNone"
    )
    assert missing_dyes == set()

    result = tool.export_workbook(input_path, output_path, mapping_path)
    assert result.row_count == 1
    assert result.missing_items == ("=1+1",)
    assert result.missing_dyes == ()

    output = load_workbook(output_path, data_only=False)
    sheet = output.active
    assert sheet.max_row == 2
    assert sheet.max_column == 8
    assert [sheet.cell(1, column).value for column in range(1, 9)] == list(tool.OUTPUT_HEADERS)
    assert sheet.cell(2, 4).value == (
        "ボルトライズ・ハット\nディアンドルロングスカート\n'=1+1"
    )
    assert sheet.cell(2, 5).value == "ジェットブラック | 染色無し\n無し"
    assert sheet.cell(2, 6).value == (
        "Boltrise Hat\nDirndl's Long Skirt\n'=1+1"
    )
    assert sheet.cell(2, 7).value == "Jet Black | No Color\nNone"
    assert sheet.cell(2, 8).value.startswith("ボルトライズ・ハット\nBoltrise Hat")
    assert sheet.cell(2, 4).alignment.wrap_text is True
    assert sheet.column_dimensions["A"].width == 40
    output.close()

print("batch loadout export ok")
